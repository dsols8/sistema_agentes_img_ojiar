# src/train_llm.py
import os
import json
from pathlib import Path
from dotenv import load_dotenv
from openai import OpenAI
import openpyxl

# Carga la clave de OpenAI desde .env
dotenv_loaded = load_dotenv()
if not dotenv_loaded:
    print("⚠️ No se cargó .env o falta OPENAI_API_KEY")
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Raíz del proyecto
ROOT = Path(__file__).resolve().parent.parent


def train_llm(excel_dir: str | Path, fixed_filename: str, model: str = "gpt-4o-2024-08-06"):
    """
    Compara el Excel original vs corregido y lanza un fine-tune usando la API de OpenAI.

    :param excel_dir: Carpeta donde están los Excel (p. ej. "excel/Estilos").
    :param fixed_filename: Nombre del Excel corregido, p.ej. "estilos_arreglado.xlsx".
    :param model: Modelo base para fine-tune. Default 'gpt-4o-mini'.
    """
    # Resuelve rutas absolutas
    ex_dir = Path(excel_dir)
    if not ex_dir.is_absolute():
        ex_dir = ROOT / excel_dir
    if not ex_dir.exists():
        raise FileNotFoundError(f"Directorio no existe: {ex_dir}")

    # Archivos original y corregido
    fixed_path = ex_dir / fixed_filename
    if not fixed_path.exists():
        raise FileNotFoundError(f"Archivo corregido no encontrado: {fixed_path}")
    base = fixed_path.stem
    orig_base = base.removesuffix("_arreglado")
    orig_path = ex_dir / f"{orig_base}.xlsx"
    if not orig_path.exists():
        raise FileNotFoundError(f"Archivo original no encontrado: {orig_path}")

    # Carga ambos Excel
    wb_orig = openpyxl.load_workbook(orig_path)
    ws_orig = wb_orig.active
    wb_fix = openpyxl.load_workbook(fixed_path)
    ws_fix = wb_fix.active

    # Encabezados de columnas
    headers = [cell.value for cell in ws_orig[1]]
    examples = []
    # Recorre filas (omitiendo encabezado)
    for row_idx in range(2, max(ws_orig.max_row, ws_fix.max_row) + 1):
        orig_vals = [ws_orig.cell(row_idx, col).value or "" for col in range(1, len(headers) + 1)]
        fix_vals = [ws_fix.cell(row_idx, col).value or "" for col in range(1, len(headers) + 1)]
        if orig_vals != fix_vals:
            prompt = json.dumps(dict(zip(headers, orig_vals)), ensure_ascii=False)
            completion = json.dumps({
                "nombre": fix_vals[headers.index("Nombre")] if "Nombre" in headers else "",
                "codigo": fix_vals[headers.index("Código")] if "Código" in headers else "",
                "precio": fix_vals[headers.index("Precio")] if "Precio" in headers else ""
            }, ensure_ascii=False)
            examples.append({"prompt": prompt, "completion": completion})

    if not examples:
        print("⚠️ No hay diferencias: nada que entrenar.")
        return

    # Guarda ejemplos en JSONL
    ft_dir = ROOT / "fine_tune"
    ft_dir.mkdir(exist_ok=True)
    jsonl_path = ft_dir / f"{orig_base}_train.jsonl"
    with open(jsonl_path, "w", encoding="utf-8") as f:
        for ex in examples:
            f.write(json.dumps(ex, ensure_ascii=False) + "\n")
    print(f"📦 JSONL listo con {len(examples)} ejemplos en {jsonl_path}")

    # Sube el archivo para fine-tune
    print("🚀 Subiendo archivo para fine-tune...")
    upload = client.files.create(
        file=open(jsonl_path, "rb"),
        purpose="fine-tune"
    )
    file_id = upload.id  # accede como atributo, no como dict
    print(f"✅ Archivo subido, file_id={file_id}")

    # Inicia fine-tune job
    print(f"🎯 Iniciando fine-tune con modelo base {model}...")
    ft_job = client.fine_tuning.jobs.create(
        training_file=file_id,
        model=model
    )
    print(f"🎉 Fine-tune iniciado, job_id={ft_job.id}")
    return ft_job
